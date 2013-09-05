"""
Module parse to/from Excel
"""

#----------------------------------------------------------------------
# ExcelFile class

import datetime
import numpy as np

from pandas.io.parsers import TextParser
from pandas.tseries.period import Period
from pandas import json
from pandas.compat import map, zip, reduce, range, lrange
from pandas.core import config
import pandas.compat as compat
from warnings import warn

# Set up the io.excel specific configuration.
writer_engine_doc = """
: string
    The default Excel engine. The options are 'openpyxl' (the default), 'xlwt'
    and 'xlsxwriter'.
"""
with config.config_prefix('io.excel'):
    config.register_option('writer_engine', None, writer_engine_doc,
                            validator=str)


def read_excel(path_or_buf, sheetname, **kwds):
    """Read an Excel table into a pandas DataFrame

    Parameters
    ----------
    sheetname : string
         Name of Excel sheet
    header : int, default 0
         Row to use for the column labels of the parsed DataFrame
    skiprows : list-like
        Rows to skip at the beginning (0-indexed)
    skip_footer : int, default 0
        Rows at the end to skip (0-indexed)
    index_col : int, default None
        Column to use as the row labels of the DataFrame. Pass None if
        there is no such column
    parse_cols : int or list, default None
        * If None then parse all columns,
        * If int then indicates last column to be parsed
        * If list of ints then indicates list of column numbers to be parsed
        * If string then indicates comma separated list of column names and
          column ranges (e.g. "A:E" or "A,C,E:F")
    na_values : list-like, default None
        List of additional strings to recognize as NA/NaN
    keep_default_na : bool, default True
        If na_values are specified and keep_default_na is False the default NaN
        values are overridden, otherwise they're appended to
    verbose : boolean, default False
        Indicate number of NA values placed in non-numeric columns

    Returns
    -------
    parsed : DataFrame
        DataFrame from the passed in Excel file
    """
    if 'kind' in kwds:
        kwds.pop('kind')
        warn("kind keyword is no longer supported in read_excel and may be "
             "removed in a future version", FutureWarning)
    return ExcelFile(path_or_buf).parse(sheetname=sheetname, **kwds)


class ExcelFile(object):
    """
    Class for parsing tabular excel sheets into DataFrame objects.
    Uses xlrd. See ExcelFile.parse for more documentation

    Parameters
    ----------
    path : string or file-like object
        Path to xls or xlsx file
    """
    def __init__(self, path_or_buf, **kwds):

        import xlrd  # throw an ImportError if we need to

        ver = tuple(map(int, xlrd.__VERSION__.split(".")[:2]))
        if ver < (0, 9):  # pragma: no cover
            raise ImportError("pandas requires xlrd >= 0.9.0 for excel "
                              "support, current version " + xlrd.__VERSION__)

        self.path_or_buf = path_or_buf
        self.tmpfile = None

        if isinstance(path_or_buf, compat.string_types):
            self.book = xlrd.open_workbook(path_or_buf)
        else:
            data = path_or_buf.read()
            self.book = xlrd.open_workbook(file_contents=data)

    def parse(self, sheetname, header=0, skiprows=None, skip_footer=0,
              index_col=None, parse_cols=None, parse_dates=False,
              date_parser=None, na_values=None, thousands=None, chunksize=None,
              **kwds):
        """Read an Excel table into DataFrame

        Parameters
        ----------
        sheetname : string or integer
            Name of Excel sheet or the page number of the sheet
        header : int, default 0
            Row to use for the column labels of the parsed DataFrame
        skiprows : list-like
            Rows to skip at the beginning (0-indexed)
        skip_footer : int, default 0
            Rows at the end to skip (0-indexed)
        index_col : int, default None
            Column to use as the row labels of the DataFrame. Pass None if
            there is no such column
        parse_cols : int or list, default None
            * If None then parse all columns
            * If int then indicates last column to be parsed
            * If list of ints then indicates list of column numbers to be
              parsed
            * If string then indicates comma separated list of column names and
              column ranges (e.g. "A:E" or "A,C,E:F")
        na_values : list-like, default None
            List of additional strings to recognize as NA/NaN
        keep_default_na : bool, default True
            If na_values are specified and keep_default_na is False the default
            NaN values are overridden, otherwise they're appended to
        verbose : boolean, default False
            Indicate number of NA values placed in non-numeric columns

        Returns
        -------
        parsed : DataFrame
            DataFrame parsed from the Excel file
        """
        has_index_names = False  # removed as new argument of API function

        skipfooter = kwds.pop('skipfooter', None)
        if skipfooter is not None:
            skip_footer = skipfooter

        return self._parse_excel(sheetname, header=header, skiprows=skiprows,
                                 index_col=index_col,
                                 has_index_names=has_index_names,
                                 parse_cols=parse_cols,
                                 parse_dates=parse_dates,
                                 date_parser=date_parser, na_values=na_values,
                                 thousands=thousands, chunksize=chunksize,
                                 skip_footer=skip_footer, **kwds)

    def _should_parse(self, i, parse_cols):

        def _range2cols(areas):
            """
            Convert comma separated list of column names and column ranges to a
            list of 0-based column indexes.

            >>> _range2cols('A:E')
            [0, 1, 2, 3, 4]
            >>> _range2cols('A,C,Z:AB')
            [0, 2, 25, 26, 27]
            """
            def _excel2num(x):
                "Convert Excel column name like 'AB' to 0-based column index"
                return reduce(lambda s, a: s * 26 + ord(a) - ord('A') + 1,
                              x.upper().strip(), 0) - 1

            cols = []
            for rng in areas.split(','):
                if ':' in rng:
                    rng = rng.split(':')
                    cols += lrange(_excel2num(rng[0]), _excel2num(rng[1]) + 1)
                else:
                    cols.append(_excel2num(rng))
            return cols

        if isinstance(parse_cols, int):
            return i <= parse_cols
        elif isinstance(parse_cols, compat.string_types):
            return i in _range2cols(parse_cols)
        else:
            return i in parse_cols

    def _parse_excel(self, sheetname, header=0, skiprows=None, skip_footer=0,
                     index_col=None, has_index_names=None, parse_cols=None,
                     parse_dates=False, date_parser=None, na_values=None,
                     thousands=None, chunksize=None, **kwds):
        from xlrd import (xldate_as_tuple, XL_CELL_DATE,
                          XL_CELL_ERROR, XL_CELL_BOOLEAN)

        datemode = self.book.datemode
        if isinstance(sheetname, compat.string_types):
            sheet = self.book.sheet_by_name(sheetname)
        else:  # assume an integer if not a string
            sheet = self.book.sheet_by_index(sheetname)

        data = []
        should_parse = {}
        for i in range(sheet.nrows):
            row = []
            for j, (value, typ) in enumerate(zip(sheet.row_values(i),
                                                 sheet.row_types(i))):
                if parse_cols is not None and j not in should_parse:
                    should_parse[j] = self._should_parse(j, parse_cols)

                if parse_cols is None or should_parse[j]:
                    if typ == XL_CELL_DATE:
                        dt = xldate_as_tuple(value, datemode)
                        # how to produce this first case?
                        if dt[0] < datetime.MINYEAR:  # pragma: no cover
                            value = datetime.time(*dt[3:])
                        else:
                            value = datetime.datetime(*dt)
                    elif typ == XL_CELL_ERROR:
                        value = np.nan
                    elif typ == XL_CELL_BOOLEAN:
                        value = bool(value)
                    row.append(value)

            data.append(row)

        if header is not None:
            data[header] = _trim_excel_header(data[header])

        parser = TextParser(data, header=header, index_col=index_col,
                            has_index_names=has_index_names,
                            na_values=na_values,
                            thousands=thousands,
                            parse_dates=parse_dates,
                            date_parser=date_parser,
                            skiprows=skiprows,
                            skip_footer=skip_footer,
                            chunksize=chunksize,
                            **kwds)

        return parser.read()

    @property
    def sheet_names(self):
        return self.book.sheet_names()


def _trim_excel_header(row):
    # trim header row so auto-index inference works
    # xlrd uses '' , openpyxl None
    while len(row) > 0 and (row[0] == '' or row[0] is None):
        row = row[1:]
    return row


class CellStyleConverter(object):
    """
    Utility Class which converts a style dict to xlrd or openpyxl style
    """

    @staticmethod
    def to_xls(style_dict, num_format_str=None):
        """
        converts a style_dict to an xlwt style object
        Parameters
        ----------
        style_dict: style dictionary to convert
        """
        import xlwt

        def style_to_xlwt(item, firstlevel=True, field_sep=',', line_sep=';'):
            """helper which recursively generate an xlwt easy style string
            for example:

              hstyle = {"font": {"bold": True},
              "border": {"top": "thin",
                        "right": "thin",
                        "bottom": "thin",
                        "left": "thin"},
              "align": {"horiz": "center"}}
              will be converted to
              font: bold on; \
                      border: top thin, right thin, bottom thin, left thin; \
                      align: horiz center;
            """
            if hasattr(item, 'items'):
                if firstlevel:
                    it = ["%s: %s" % (key, style_to_xlwt(value, False))
                          for key, value in item.items()]
                    out = "%s " % (line_sep).join(it)
                    return out
                else:
                    it = ["%s %s" % (key, style_to_xlwt(value, False))
                          for key, value in item.items()]
                    out = "%s " % (field_sep).join(it)
                    return out
            else:
                item = "%s" % item
                item = item.replace("True", "on")
                item = item.replace("False", "off")
                return item

        if style_dict:
            xlwt_stylestr = style_to_xlwt(style_dict)
            style = xlwt.easyxf(xlwt_stylestr, field_sep=',', line_sep=';')
        else:
            style = xlwt.XFStyle()
        if num_format_str is not None:
            style.num_format_str = num_format_str

        return style

    @staticmethod
    def to_xlsx(style_dict):
        """
        converts a style_dict to an openpyxl style object
        Parameters
        ----------
        style_dict: style dictionary to convert
        """

        from openpyxl.style import Style
        xls_style = Style()
        for key, value in style_dict.items():
            for nk, nv in value.items():
                if key == "borders":
                    (xls_style.borders.__getattribute__(nk)
                     .__setattr__('border_style', nv))
                else:
                    xls_style.__getattribute__(key).__setattr__(nk, nv)

        return xls_style

    @staticmethod
    def to_xlsxwriter(workbook, style_dict, num_format_str=None):
        """
        Converts a style_dict to an XlxsWriter format object.
        Parameters
        ----------
        workbook:   Reference to the ExcelWriter XlxsWriter workbook.
        style_dict: Style dictionary to convert.
        num_format: Optional number format for the cell format.
        """
        if style_dict is None:
            return None

        # Create a XlsxWriter format object.
        xl_format = workbook.add_format()

        # Map the cell font to XlsxWriter font properties.
        if style_dict.get('font'):
            font = style_dict['font']
            if font.get('bold'):
                xl_format.set_bold()

        # Map the cell borders to XlsxWriter border properties.
        if style_dict.get('borders'):
            xl_format.set_border()

        if num_format_str is not None:
            xl_format.set_num_format(num_format_str)

        return xl_format


def _conv_value(val):
    # convert value for excel dump
    if isinstance(val, np.int64):
        val = int(val)
    elif isinstance(val, np.bool8):
        val = bool(val)
    elif isinstance(val, Period):
        val = "%s" % val

    return val


class ExcelWriter(object):
    """
    Class for writing DataFrame objects into excel sheets, uses xlwt for xls,
    openpyxl for xlsx.  See DataFrame.to_excel for typical usage.

    Parameters
    ----------
    path : string
        Path to xls file
    """
    def __init__(self, path, engine=None, **engine_kwargs):

        if engine is None:
            default = config.get_option('io.excel.writer_engine')
            if default is not None:
                engine = default
            elif path.endswith('.xls'):
                engine = 'xlwt'
            else:
                engine = 'openpyxl'

        try:
            writer_init = getattr(self, "_init_%s" % engine)
        except AttributeError:
            raise ValueError("No engine: %s" % engine)

        writer_init(path, **engine_kwargs)

        self.sheets = {}
        self.cur_sheet = None

    def save(self):
        """
        Save workbook to disk
        """
        if self.engine == 'xlsxwriter':
            self.book.close()
        else:
            self.book.save(self.path)

    def write_cells(self, cells, sheet_name=None, startrow=0, startcol=0):
        """
        Write given formated cells into Excel an excel sheet

        Parameters
        ----------
        cells : generator
            cell of formated data to save to Excel sheet
        sheet_name : string, default None
            Name of Excel sheet, if None, then use self.cur_sheet
        startrow: upper left cell row to dump data frame
        startcol: upper left cell column to dump data frame
        """
        if sheet_name is None:
            sheet_name = self.cur_sheet

        if sheet_name is None:  # pragma: no cover
            raise ValueError('Must pass explicit sheet_name or set '
                            'cur_sheet property')

        try:
            _writecells = getattr(self, "_writecells_%s" % self.engine)
        except AttributeError:
            raise ValueError("No _writecells_%s() method" % self.engine)

        _writecells(cells, sheet_name, startrow, startcol)

    def _writecells_openpyxl(self, cells, sheet_name, startrow, startcol):
        # Write the frame cells using openpyxl.
        from openpyxl.cell import get_column_letter

        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = self.book.create_sheet()
            wks.title = sheet_name
            self.sheets[sheet_name] = wks

        for cell in cells:
            colletter = get_column_letter(startcol + cell.col + 1)
            xcell = wks.cell("%s%s" % (colletter, startrow + cell.row + 1))
            xcell.value = _conv_value(cell.val)
            if cell.style:
                style = CellStyleConverter.to_xlsx(cell.style)
                for field in style.__fields__:
                    xcell.style.__setattr__(field,
                                            style.__getattribute__(field))

            if isinstance(cell.val, datetime.datetime):
                xcell.style.number_format.format_code = "YYYY-MM-DD HH:MM:SS"
            elif isinstance(cell.val, datetime.date):
                xcell.style.number_format.format_code = "YYYY-MM-DD"

            # merging requires openpyxl latest (works on 1.6.1)
            # todo add version check
            if cell.mergestart is not None and cell.mergeend is not None:
                cletterstart = get_column_letter(startcol + cell.col + 1)
                cletterend = get_column_letter(startcol + cell.mergeend + 1)

                wks.merge_cells('%s%s:%s%s' % (cletterstart,
                                               startrow + cell.row + 1,
                                               cletterend,
                                               startrow + cell.mergestart + 1))

    def _writecells_xlwt(self, cells, sheet_name, startrow, startcol):
        # Write the frame cells using xlwt.
        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = self.book.add_sheet(sheet_name)
            self.sheets[sheet_name] = wks

        style_dict = {}

        for cell in cells:
            val = _conv_value(cell.val)

            num_format_str = None
            if isinstance(cell.val, datetime.datetime):
                num_format_str = "YYYY-MM-DD HH:MM:SS"
            if isinstance(cell.val, datetime.date):
                num_format_str = "YYYY-MM-DD"

            stylekey = json.dumps(cell.style)
            if num_format_str:
                stylekey += num_format_str

            if stylekey in style_dict:
                style = style_dict[stylekey]
            else:
                style = CellStyleConverter.to_xls(cell.style, num_format_str)
                style_dict[stylekey] = style

            if cell.mergestart is not None and cell.mergeend is not None:
                wks.write_merge(startrow + cell.row,
                                startrow + cell.mergestart,
                                startcol + cell.col,
                                startcol + cell.mergeend,
                                val, style)
            else:
                wks.write(startrow + cell.row,
                          startcol + cell.col,
                          val, style)

    def _writecells_xlsxwriter(self, cells, sheet_name, startrow, startcol):
        # Write the frame cells using xlsxwriter.
        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = self.book.add_worksheet(sheet_name)
            self.sheets[sheet_name] = wks

        style_dict = {}

        for cell in cells:
            val = _conv_value(cell.val)

            num_format_str = None
            if isinstance(cell.val, datetime.datetime):
                num_format_str = "YYYY-MM-DD HH:MM:SS"
            if isinstance(cell.val, datetime.date):
                num_format_str = "YYYY-MM-DD"

            stylekey = json.dumps(cell.style)
            if num_format_str:
                stylekey += num_format_str

            if stylekey in style_dict:
                style = style_dict[stylekey]
            else:
                style = CellStyleConverter.to_xlsxwriter(self.book,
                                                         cell.style,
                                                         num_format_str)
                style_dict[stylekey] = style

            if cell.mergestart is not None and cell.mergeend is not None:
                wks.merge_range(startrow + cell.row,
                                startrow + cell.mergestart,
                                startcol + cell.col,
                                startcol + cell.mergeend,
                                val, style)
            else:
                wks.write(startrow + cell.row,
                          startcol + cell.col,
                          val, style)

    def _init_xlwt(self, filename, **engine_kwargs):
        # Use the xlwt module as the Excel writer.
        import xlwt

        self.engine = 'xlwt'
        self.path = filename
        self.book = xlwt.Workbook()
        self.fm_datetime = xlwt.easyxf(num_format_str='YYYY-MM-DD HH:MM:SS')
        self.fm_date = xlwt.easyxf(num_format_str='YYYY-MM-DD')

    def _init_openpyxl(self, filename, **engine_kwargs):
        # Use the openpyxl module as the Excel writer.
        from openpyxl.workbook import Workbook

        self.engine = 'openpyxl'
        self.path = filename
        # Create workbook object with default optimized_write=True.
        self.book = Workbook()
        # Openpyxl 1.6.1 adds a dummy sheet. We remove it.
        if self.book.worksheets:
            self.book.remove_sheet(self.book.worksheets[0])

    def _init_xlsxwriter(self, filename, **engine_kwargs):
        # Use the xlsxwriter module as the Excel writer.
        import xlsxwriter

        options = dict(engine_kwargs)

        options.setdefault('default_date_format', 'YYYY-MM-DD HH:MM:SS')

        self.engine = 'xlsxwriter'
        self.book = xlsxwriter.Workbook(filename, options)
